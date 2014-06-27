# -*- coding: utf-8 -*-
#
#	Clase para la importacion de facturas comerciales
#	en formato XLSX
#

from openpyxl import load_workbook
import xlrd
from conector_mysql import Conexion


# verificar si es XLS o XLSX para saber el objeto a utilizar.
"""
class Mercancia_Transito():
	
	dict_mercancia_ = {
		'guia': ''
		,'factura': ''
		,'codigo_producto': ''
		,'precio_unitario':0
		,'cantidad': 0
		,'unidad_medida':''
		,'importe':''
		,'clave_proveedor':''
		,'nombre_proveedor':''
		,'pais_origen':''
		,'incoterm': ''
		,'fecha_embarque':''
		,'pais_embarque':''
		,'fecha_estimada':''
		,'fecha_arribo':''
		,'moneda': ''
		,'fraccion': ''
		,'iva': 00.00
		,'igi': 00.00}
	facturas_ = []
	wb_ = load_workbook(filename= 'C:\\Users\\AlfredoVG.DAIMEX\\Documents\\Codigo\\Proyecto Carga de Facturas\\Reporte de mercancía en transito a detalle.xlsx', use_iterators= True)
	ws_ = wb_.worksheets[0]
	ban_ = 0
	indices_ = {
		'House BL No.':None
		,'Invoice Number': None
		,'Material Code': None
		,'Net Amount':None
		,'Delivery Qty.': None
		,'Qty Unit':None
		,' Total Amount':None
		,'clave_proveedor':None
		,'nombre_proveedor':None
		,'pais_origen':None
		,'Incoterms1': None
		,'Loading Port Out':None
		,'pais_embarque':None
		,'Port ETA':None
		,'fecha_arribo':None
		,'Currency': None
	}
	for row in ws_.iter_rows():
		if ban_ == 0: #row[0].internal_value=='Delivery No.':
			for cell_ in row:
				
			
		else:
			mercancia_ = dict_mercancia_.copy()
			mercancia_['guia']=row[2].internal_value
			mercancia_['factura'] = row[3].internal_value
			mercancia_['codigo_producto'] =row[5].internal_value
			mercancia_['precio_unitario'] =row[14].internal_value
			mercancia_['cantidad'] =row[6].internal_value
			mercancia_['unidad_medida'] =row[13].internal_value
			mercancia_['importe'] =row[15].internal_value
			mercancia_['incortem'] = row[17].internal_value
			mercancia_['moneda'] = row[18].internal_value
			mercancia_['clave_proveedor'] =row[24].internal_value
			mercancia_['nombre_proveedor'] =row[25].internal_value
			mercancia_['pais_origen'] =	row[28].internal_value
			facturas_.append(mercancia_)

	
	def __init__(self):
		
		
		return
"""
	


class Importa_Factura:
	_archivo = ''
	_libro = ''
	_hoja = ''
	_factura = ''
	_hojas_ = ''
	_filas = -1
	_columas = -1
	_cantidad = None
	_descripcion = None
	_importe = None
	_precio_unitario = None
	_cliente = 0
	_proveedor = 0
	_patente = 0
	_conexion = None
	_estructura_factura = {
		'numero_factura':''
		,'fecha_factura':''
		,'proveedor':{'nombre':'','direccion':''}
		,'cliente':{'nombre':'','direccion':''}
		,'pais_origen': ''
		,'mercancias':[]
		, 'moneda':''
	}
	
	def __init__(self, _archivo , _clave_proveedor = ''
				 , _clave_cliente = '', _patente = ''):
		'''
		Constructor recibe por defecto la ruta del archivo
		'''
		self._archivo = _archivo
		self._cliente = _clave_cliente
		self._proveedor = _clave_proveedor
		self._patente = _patente
		self.conexion()
		return

	def serializa_xls(self):
		'''
		realiza el proceso de serializacion
		'''
		self._libro = xlrd.open_workbook(self._archivo)
		self._factura = self._libro.sheet_by_name(self._libro.sheet_names()[0])
		self._filas = self._factura.nrows -1
		self._columnas = self._factura.ncols -1

		# variables de control, para ubicar la columna  en la cual se encuentran
		# estos valores
		descripcion_ = None
		cantidad_ = None
		precio_unitario_ = None
		importe_ = None

		# fila_
		fila_ = -1
		ban_cuerpo_ = False

		while fila_ < self._filas:
			fila_ +=1
			valor_fila_ = self._factura.row(fila_)
			columna_ = -1
			while columna_ < self._columnas:
				columna_ +=1
				tipo_celda_ = self._factura.cell_type(fila_,columna_)
				valor_celda_ = self._factura.cell_value(fila_,columna_)
				try:
					if valor_celda_.strip() == 'Seller':
						# en caso de encontrar la celda con Seller se 
						# invoca el metodo correspondiente para obtener la informacion
						# del vendedor que se encuentra debajo de este tag y ante sdel tag Consignee 
						self._estructura_factura['proveedor'] = self.get_proveedor(fila_, columna_)
					elif valor_celda_.strip() == 'Invoice No. & Date':
						self._estructura_factura['numero_factura'], self._estructura_factura['fecha_factura'] = self.get_numero_factura(fila_,columna_)
					elif valor_celda_.strip() == 'Buyer' :
						self._estructura_factura['cliente'] = self.get_cliente(fila_,columna_)
					elif valor_celda_.strip() == 'Country of Origin' :
						self._estructura_factura['pais_origen'] = self.get_pais_origen(fila_,columna_)
					elif valor_celda_.strip() == 'Good Description' :
						self._descripcion = columna_
					elif valor_celda_.strip() == 'Quantity' :
						self._cantidad = columna_
					elif valor_celda_.strip() == 'Unit Pirce':
						self._precio_unitario = columna_
					elif valor_celda_.strip() == 'Amount' :
						self._importe = columna_
				except AttributeError: 
					pass

				if self._descripcion != None and self._cantidad != None and self._precio_unitario != None and self._importe != None :
					ban_cuerpo_ = True
					break
			if ban_cuerpo_ :
				break
		self.get_mercancias(fila_)

		return
	
	def get_estructura(self):
		return self._estructura_factura

	def get_pais_origen(self,_fila,_columna,_tipo='xls'):
		if _tipo == 'xls':
			pais_ = self._factura.cell_value(_fila+1,_columna)
		else:
			pais_ = None
		return pais_

	def get_numero_factura(self, _fila , _columna,_tipo='xls' ):
		if _tipo == 'xls':

			numero_   = self._factura.cell_value(_fila+1,_columna).split(' ')[0]
			fecha_ = self._factura.cell_value(_fila+1,_columna).split(' ')[-1]
		else:
			numero_ , fecha_ = None 
		return numero_ , fecha_

	def get_proveedor(self,_fila , _columna,_tipo='xls'):
		if _tipo == 'xls':
			nombre_ = self._factura.cell_value(_fila+1,_columna)
			direccion_ = '%s %s %s %s %s'%(self._factura.cell_value(_fila+2,_columna),self._factura.cell_value(_fila+3,_columna), self._factura.cell_value(_fila+4,_columna), self._factura.cell_value(_fila+5,_columna), self._factura.cell_value(_fila+6,_columna))
		else:
			nombre_,direccion_ = None
		return nombre_, direccion_

	def get_cliente(self,_fila , _columna,_tipo='xls'):
		if _tipo == 'xls':
			nombre_ = '%s %s'%(self._factura.cell_value(_fila+1,_columna), self._factura.cell_value(_fila+2,_columna))
			direccion_ = '%s %s %s %s'%(self._factura.cell_value(_fila+3,_columna), self._factura.cell_value(_fila+4,_columna), self._factura.cell_value(_fila+5,_columna), self._factura.cell_value(_fila+6,_columna))
		else:
			nombre_, direccion_ = None

		return nombre_, direccion_

	def get_mercancias(self,_fila ,_tipo='xls'):
		if _tipo == 'xls':
			ban_renglon_ = 0
			partida_ = 0
			mercancia_ = {
				'partida':''
				,'codigo_producto':''
				,'cantidad':''
				,'descripcion':''
				,'costo_unitario':''
				,'importe':''
				,'especificaciones':''
				,'unidad':''
				,'existe_catalogo': False
			}
			fila_ = _fila
			while fila_ < self._filas:
				fila_ +=1
				if ban_renglon_ ==0: # renglon donde se encuentra la descripcion
					
					if self._factura.cell_value(fila_, self._descripcion) != 'TOTAL':
						partida_ +=1

						mercancia_['descripcion'] = self._factura.cell_value(fila_, self._descripcion)
						mercancia_['partida'] = partida_
					else:
						return

				elif ban_renglon_ ==1: # renglon donde se encuentra la Clave, precio y unidad
					mercancia_['codigo_producto'] = self._factura.cell_value(fila_, self._descripcion)
					mercancia_['cantidad'] = self._factura.cell_value(fila_, self._cantidad).split(' ')[0]
					mercancia_['unidad'] = self._factura.cell_value(fila_, self._cantidad).split(' ')[-1]
					mercancia_['costo_unitario']= self._factura.cell_value(fila_, self._precio_unitario).split(' ')[0]
					mercancia_['importe'] = self._factura.cell_value(fila_, self._importe).split(' ')[0]
					self._estructura_factura['moneda'] = self._factura.cell_value(fila_, self._importe).split(' ')[-1]

				elif ban_renglon_ ==2: # renglon donde se encuentran especificaciones
					mercancia_['especificaciones'] = self._factura.cell_value(fila_, self._descripcion)
					self._estructura_factura['mercancias'].append(mercancia_)

				if ban_renglon_ < 2:
					ban_renglon_ +=1
				else:
					ban_renglon_ = 0
					mercancia_ = {
						'partida':''
						,'codigo_producto':''
						,'cantidad':''
						,'descripcion':''
						,'costo_unitario':''
						,'importe':''
						,'especificaciones':''
						,'unidad':''
					}

		return

	def conexion(self):
		self._conexion = Conexion()
		return

	def verifica_factura(self):
		'''
		Metodo que realiza la verificacion de la factura,
		es decir, saber si esta factura para la clave de cliente
		y proveedor ya fue registrada anteriormente.
		'''
		
		consulta_ = """
			select n_cve_fact
			from enc001_facturas
			where t_numfac = '%(factura_)s'
			and n_patente = %(patente_)s
			and i_cve_cli = %(cliente_)s
			and i_cve_pro = %(proveedor_)s
					""" %{'factura_': self._estructura_factura['numero_factura']
						  , 'patente_':self._patente
						  ,'cliente_':self._cliente
						  ,'proveedor_':self._proveedor}
		verificacion_ =self._conexion.get_resultados(consulta_)
		if len(verificacion_)>0:
			return verificacion_[0]['n_cve_fact']
		else:
			return 0

	def verifica_mercancia(self):
		"""
		Funcion que busca entre el catalogo de mercancias revisando la informacion
		e indicando si se encuentra o no.
		"""
		
		for n_mercancia_ in range(len(self._estructura_factura['mercancias'])):
			consulta_ = """
				select frac05, f8va05, desc05e
				from  cat005_artic 
				where cpro05 = '%(clave_producto_)s'
				and clie05 = %(cliente_)s
				and prov05 = %(proveedor_)s
				"""%{'clave_producto_': self._estructura_factura['mercancias'][n_mercancia_]['codigo_producto']
					, 'cliente_': self._cliente
					, 'proveedor_': self._proveedor}
			fracciones_ = self._conexion.get_resultados(consulta_)
			if len(fracciones_)>0:
				self._estructura_factura['mercancias'][n_mercancia_]['existe'] = True
				self._estructura_factura['mercancias'][n_mercancia_].update(fracciones_[0])
				self._estructura_factura['mercancias'][n_mercancia_].update({'class':'success'})
			else:
				self._estructura_factura['mercancias'][n_mercancia_].update({'class':'danger'})
		
		return
	


#if __name__ == '__main__':
#	xls_ = Importa_Factura('C:\\Users\\AlfredoVG.DAIMEX\\Documents\\Codigo\\Proyecto Carga de Facturas\\C6H0_SR_9006569384.xls')
#	factura_ = xls_.serializa_xls()
#	print factura_
