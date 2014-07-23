# -*- coding: utf-8 -*-
from utils import conector_mysql
import xlwt, openpyxl as xlsx, csv, operator
from os import path
from openpyxl.styles.fonts import Font
from openpyxl.styles.colors import Color
from openpyxl.styles import Style



class Reporte_vivas:
    _conexion = None
    _columnas = ['Tipo Operacion','Referencia','Pedimento','Cliente','Clave Pedimento',
                      'Fecha Entrada','Arribo Estimado','Alta de Referencia','Días a la fecha','Estatus','Ejecutivos']
    _fila_actual = 0
    _titulo = 'Referencias Vivas'
    _directorio = None
    _reporte = None
    _libro = None
    def __init__(self,_directorio):
        
        self._directorio = _directorio
        
        return
    
    def genera_xlsx(self):
        
        self._libro = xlsx.Workbook()        
        self._reporte = self._libro.active
        self._reporte.title = self._titulo
        
        self.llena_encabezado('xlsx')
        self.llena_cuerpo('xlsx',self.consulta())
        #ruta_ = '%s'%path.join(self._directorio,self._titulo)
        
        #self._libro.save(ruta_)
        self._conexion.__close__()
        return self._libro
        
    
    def consulta (self):
        self._conexion = conector_mysql.Conexion()
        #seleccionando la base de datos.
        self._conexion.exe("use dai_extranet")
        # Estableciendo tamaño de la cadena para concatenar los estatos y ejecutivos
        self._conexion.exe("set session group_concat_max_len = 1569325555")
        # Estableciendo variable para la fecha de hoy y aprovechar el sistema de caché.
        self._conexion.exe("set @hoy_ = curdate()")
        
        #revisando variables para generar condiciones adicionales.
        
        #nombre de cliente
        
        #fecha de pago
        
        #fecha de alta
        
        
        #definiendo consulta principal.
        base_sql_ = """
            select 'impo' as "Tipo Operacion"
                , op_.refcia01 as "Referencia"
                , op_.numped01 as Pedimento
                , op_.nomcli01 as Cliente
                , op_.cveped01  as "Clave Pedimento"
                , if(op_.fecent01>0,date_format(op_.fecent01,'%d-%m-%Y'),'') as "Fecha Entrada"
                , if(ref_.feta01>0,date_format(ref_.feta01,'%d-%m-%Y'),'') as "Arribo Estimado"
                , if(ref_.frec01>0,date_format(ref_.frec01,'%d-%m-%Y'),'') as "Alta de Referencia"
                , @fecha_ := greatest(fecent01, feta01, frec01) as fecha
                , @fecha_
                , datediff(@hoy_,@fecha_) as "Días a la fecha"
                ,datediff(@hoy_,@fecha_) as diferencia
                ,ifnull(group_concat( distinct 
                    concat(
                        'Estado: ',eta_.d_nombre
                        ,' Fecha: ', date_format(f_fecha,'%d-%m-%Y')
                        , if( rtrim(ltrim(edo_.m_observ)) != '',concat(' Observ: ', rtrim(ltrim(edo_.m_observ))),'') )
                        separator '\r\n'),'') as Estatus
                , ejecutivos as Ejecutivos
            from ssdagi01 op_
            left join c01refer ref_ on op_.refcia01 = ref_.refe01
            LEFT JOIN ( select dgrp_.clie09 , group_concat(grp_.nomb08 separator ', ') as ejecutivos
            from c09cligr dgrp_
            left join c08grupo grp_ on dgrp_.grup09 = grp_.grup08
            group by dgrp_.clie09 ) as ejecutivos_ on op_.cvecli01 = ejecutivos_.clie09
            left join etxpd edo_ on edo_.c_referencia = op_.refcia01
            left join etaps eta_ on edo_.n_etapa =eta_.n_etapa
            where ref_.modo01 = 'T' 
            and (date_format(ref_.fdsp01,'%Y%m%d') = '00000000' or date_format(ref_.fdsp01,'%Y%m%d') = '' )
            and op_.cveped01 != 'R1' and ref_.csit01 != 'FIN'
            group by op_.refcia01
            
            union all 
            
            select 'expo' as tipo_pedimento
                , op_.refcia01
                , op_.numped01
                , op_.nomcli01
                , op_.cveped01 
                , if(op_.fecpre01>0,date_format(op_.fecpre01,'%d-%m-%Y'),'') as entrada
                , if(ref_.feta01>0,date_format(ref_.feta01,'%d-%m-%Y'),'') as arribo_estimado
                , if(ref_.frec01>0,date_format(ref_.frec01,'%d-%m-%Y'),'') as alta_referencia
                , @fecha_ := greatest(fecpre01, feta01, frec01) as fecha
                , @fecha_
                , datediff(@hoy_,@fecha_) as diferencia
                , datediff(@hoy_,@fecha_)
                ,ifnull(group_concat( distinct 
                    concat(
                        'Estado: ',eta_.d_nombre
                        ,' Fecha: ', date_format(f_fecha,'%d-%m-%Y')
                        , if( rtrim(ltrim(edo_.m_observ)) != '',concat(' Observ: ', rtrim(ltrim(edo_.m_observ))),'') )
                        separator '\r\n'),'') as Estatus
                , ejecutivos
            from ssdage01 op_
            left join c01refer ref_ on op_.refcia01 = ref_.refe01
            LEFT JOIN ( select dgrp_.clie09 , group_concat(grp_.nomb08 separator ', ') as ejecutivos
            from c09cligr dgrp_
            left join c08grupo grp_ on dgrp_.grup09 = grp_.grup08
            group by dgrp_.clie09 ) as ejecutivos_ on op_.cvecli01 = ejecutivos_.clie09
            left join etxpd edo_ on edo_.c_referencia = op_.refcia01
            left join etaps eta_ on edo_.n_etapa =eta_.n_etapa
            where ref_.modo01 = 'T' 
            and (date_format(ref_.fdsp01,'%Y%m%d') = '00000000' or date_format(ref_.fdsp01,'%Y%m%d') = '' )
            and op_.cveped01 != 'R1' and ref_.csit01 != 'FIN'
            group by op_.refcia01
            
            order by diferencia desc, Referencia desc
        """
        
        resultados_ = self._conexion.get_resultados(base_sql_)
        return resultados_
    
    def llena_encabezado(self,_tipo):
        
        if _tipo =='xlsx':
            for n_campo_ in range(len(self._columnas)):
                celda_ = self._reporte.cell(row = self._fila_actual, column = n_campo_)
                celda_.value = self._columnas[n_campo_]
                #celda_.styles= Style(font=)
        else:
            pass
        
        self._fila_actual += 1
        return 
    
    def llena_cuerpo(self,_tipo,_consulta):
        
        if _tipo == 'xlsx':
            #log_ = open('c:\\temp\\xlsx.txt','w')
            #log_.write(_consulta.__str__())
            for fila_ in _consulta:
                #log_.write(type(fila_))
                for n_campo_ in range(len(self._columnas)):
                    
                    #log_.write(fila_[self._columnas[n_campo_]])
                    #log_.write('\n')
                    self._reporte.cell(row = self._fila_actual, column = n_campo_).value = fila_[self._columnas[n_campo_]]
                self._fila_actual +=1
            #log_.close()
        return 