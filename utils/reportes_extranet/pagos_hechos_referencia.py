# -*- coding: utf-8 -*-
from utils import conector_mysql
import xlwt, openpyxl as xlsx, csv, operator
from os import path
import os
import re
from openpyxl.styles.fonts import Font
from openpyxl.styles.colors import Color
from openpyxl.styles import Style
import datetime


class Pagos_hechos:
    _conexion = None
    _ruta = None
    _fila_actual = 0
    _titulo = "Pagos Hechos %s"%datetime.date.today().isoformat()
    _reporte = None
    _libro = None
    _honorarios = 0
    _columnas = ['no','oficina',	'cliente',	'destinatario',	'factura_anterior',	'refcia01'
                    ,'numgui04',	'factura_comercial',	'factura_actual',	'pedimento'
                    ,'ingreso_sap',	'cuenta_americana',	'impuestos_pedimento',	'rectificacion'
                    ,'desconsolidacion','flete_terrestre' ,'flete_terrestre_blindado', 'retencion'
                    ,	'flete_aereo',	'maniobras_cargo_minimo'
                    ,'transmision_ams',	'almacenaje',	'custodia',	'previos',	'maniobras'
                    ,'manejo',	'montacargas_servicio_extraordinario',	'servicio_extraordinario'
                    ,'montacargas',	'pagos_hechos',	'importe_honorarios',	'iva_cg',	'anticipos', 'total'
                    ,'fecha_pago_pedimento',	'tipo_mercancia']
    _referencias= []
    #_defalut_storage = None
    
    def __init__(self, _ruta, _referencias , _honorarios = 1200):
        #self._defalut_storage = _default_storage
        self._ruta = _ruta
        self._honorarios = _honorarios
        self._referencias = _referencias
        return
    
    def genera_xlsx(self):
        self._libro = xlsx.Workbook()        
        self._reporte = self._libro.active
        self._reporte.title = self._titulo
        
        #self._titulo = '%s Ref %s'%(self._titulo,len(self._referencias))
        self.llena_encabezado('xlsx')
        self.llena_cuerpo('xlsx',self.consulta())
        ruta_ = '%s.xlsx'%path.join(self._ruta,self._titulo)
        
        if path.isfile(ruta_):
            archivos_ = os.listdir(self._ruta)
            iguales_ = re.compile(self._titulo)
            
            n_ = len(iguales_.findall(','.join(archivos_))) +1
            ruta_ = '%s (%s).xlsx'%(path.join(self._ruta,self._titulo),n_)
            archivo_ = '%s (%s).xlsx'%(self._titulo,n_)
        else:
            archivo_ = '%s.xlsx'%self._titulo
        
        self._libro.save(ruta_)
        self._conexion.__close__()
        return archivo_
    
    def llena_encabezado(self,_tipo):
        
        if _tipo =='xlsx':
            for n_campo_ in range(len(self._columnas)):
                celda_ = self._reporte.cell(row = self._fila_actual, column = n_campo_)
                celda_.value = self._columnas[n_campo_]
        else:
            pass
        
        self._fila_actual += 1
        return
    
    def llena_cuerpo(self,_tipo,_consulta):
        
        if _tipo == 'xlsx':
            for fila_ in _consulta:
                for n_campo_ in range(len(self._columnas)):
                    self._reporte.cell(row = self._fila_actual, column = n_campo_).value = fila_[self._columnas[n_campo_]]
                self._fila_actual +=1
        return
    
    def consulta(self):
        self._conexion = conector_mysql.Conexion()
        self._conexion.exe("use dai_extranet")
        self._conexion.exe("set @num_ = 0")
        self._conexion.exe("set @honorarios_ = %s"%self._honorarios)

        base_sql_ = """
            select  @num_ := @num_+1 as 'no'
                , 'DaiSC' as oficina
                , op_.nomcli01 as cliente
                , '' as destinatario
                , '' as factura_anterior
                , op_.refcia01
                , gui_.numgui04
                , group_concat( distinct fact_.numfac39 order by fact_.numfac39 desc separator '|') as 'factura_comercial'
                , '' as factura_actual
                , op_.numped01 as pedimento
                , '' as ingreso_sap
                , 0.00 as cuenta_americana
                , sum( if( epg_.conc21 = 1 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21 ,0)) as 'impuestos_pedimento'
                , sum( if( epg_.conc21 = 31 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as rectificacion
                , sum( if( epg_.conc21 = 6 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as desconsolidacion
                , sum( if( epg_.conc21 = 7 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21,0)) as flete_terrestre
                , sum( if( epg_.conc21 = 324 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21,0)) as flete_terrestre_blindado
                , sum( if( epg_.conc21 = 324 or epg_.conc21 = 7 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mfle21 *.04,0)) as retencion
                , sum( if( epg_.conc21 = 3 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21,0)) as flete_aereo
                , sum( if( epg_.conc21 = 325 ,case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as maniobras_cargo_minimo
                , sum( if( epg_.conc21 = 276 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as transmision_ams
                , sum( if( epg_.conc21 = 10 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as almacenaje	
                , sum( if( epg_.conc21 = 82 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as custodia
                , sum( if( epg_.conc21 = 102 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as previos
                , sum( if( epg_.conc21 = 127 ,case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as maniobras
                , sum( if( epg_.conc21 = 141 ,case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0))  as manejo
                , sum( if( epg_.conc21 = 329 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as montacargas_servicio_extraordinario
                , sum( if( epg_.conc21 = 63 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as servicio_extraordinario
                , sum( if( epg_.conc21 = 11 or epg_.conc21 = 326 , case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21, 0)) as montacargas
                , ifnull(sum( case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21 ),0) as pagos_hechos
                , @honorarios_ as importe_honorarios
                , @honorarios_ *.16 as iva_cg
                , ifnull(dmov_.mont11,0)  as anticipos
                , ifnull(sum( case epg_.deha21 when 'C' then -1 else 1 end  * dpg_.mont21 ),0) + @honorarios_ + (@honorarios_ *.16) - ifnull(dmov_.mont11,0) as total
                , op_.fecpag01 as fecha_pago_pedimento
                , (select group_concat(distinct art_.tpmerc05 ) from d05artic art_ where art_.refe05 = op_.refcia01) as tipo_mercancia
            from ssdagi01 op_ 
            left join c01refer ref_ on op_.refcia01 = ref_.refe01
            left join ssfact39 fact_ on op_.refcia01 = fact_.refcia39
            left join d21paghe dpg_ on dpg_.refe21 = op_.refcia01 
            left join e21paghe epg_ on dpg_.foli21 = epg_.foli21 and year(dpg_.fech21) = year(epg_.fech21) and dpg_.tmov21 = epg_.tmov21
            left join ssguia04 gui_ on op_.refcia01 = gui_.refcia04 and gui_.idngui04 = 2 
            left join d11movim dmov_ on op_.refcia01 = dmov_.refe11 and dmov_.desc11 = 'ANTICIPO'
            where op_.refcia01 in ('%s') 
            group by  op_.refcia01
            """%"','".upper().join(self._referencias)
        
        
        resultados_ = self._conexion.get_resultados(base_sql_)
        
        return resultados_