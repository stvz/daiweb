# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import xlrd, operator
from conector_mysql import Conexion
from os import path

def tipo_archivo(_archivo):
    """
    Analiza el nombre del archivo y devuelve el tipo de archivo:
    1 para excel <2003
    2 para excel >2007
    0 para cualquier otro
    """
    extension_ = path.splitext(_archivo)[1]
    
    if extension_ == '.xls':
        tipo_ = 1
    elif extension_ == '.xlsx':
        tipo_ = 2
    else:
        tipo_ = 0
    
    return tipo_

class Mercancias:
    """
    Clase para el manejo del archivos que proporciona Samsung (Reporte de mercancias en transito)
    Contiene las siguientes columnas
        'Delivery No.','Reference Doc(SO/PO)','House BL No.','Invoice Number'
        ,'Import Customs Number','Material Code','Delivery Qty.','Loading Port Out'
        ,'Port ETA','Port ATA','Import Clearance','Warehouse ATA','Mtl. Division'
        ,'Qty Unit','Net Amount','Total Amount','Currency','Incoterms1','Delivery Date'
        ,'DO Created on','SR No','BL No.','Master BL No.','Transportation Type'
        ,'GS Vendor','GS Vendor Name','PO Vendor','PO Vendor Name','Forwarder Code'
        ,'Forwarder Name','Loading Country','Loading Port','Discharging Country'
        ,'Discharging Port','Final Dest.','Country','Destination Port','Related DO'
    """
    
    _tipo= 0
    _archivo = ''
    _mercancias=[]
    _campos = {'Delivery No.':None,'Reference Doc(SO/PO)':None,'House BL No.':None,'Invoice Number':None
               ,'Import Customs Number':None,'Material Code':None,'Delivery Qty.':None,'Loading Port Out':None
               ,'Port ETA':None,'Port ATA':None,'Import Clearance':None,'Warehouse ATA':None,'Mtl. Division':None
               ,'Qty Unit':None,'Net Amount':None,'Total Amount':None,'Currency':None,'Incoterms1':None,'Delivery Date':None
               ,'DO Created on':None,'SR No':None,'BL No.':None,'Master BL No.':None,'Transportation Type':None
               ,'GS Vendor':None,'GS Vendor Name':None,'PO Vendor':None,'PO Vendor Name':None,'Forwarder Code':None
               ,'Forwarder Name':None,'Loading Country':None,'Loading Port':None,'Discharging Country':None
               ,'Discharging Port':None,'Final Dest.':None,'Country':None,'Destination Port':None,'Related DO':None
               ,'Repetida':None
            }
    def __init__(self, _archivo ):
        self._tipo = tipo_archivo(_archivo)
        
        if self._tipo != 0:
            self._archivo = _archivo
            self.carga_archivo()
        else:
            raise "El tipo de archivo o extension, no coinciden con los de excel."
    
    def carga_archivo(self):
        
        if self._tipo == 1:
            estatus_, mercancia_ = self.importa_xls()
        elif self._tipo == 2:
            estatus_, mercancia_ = self.importa_xlsx()
        else:
            estatus_ = ''
        
        return estatus_, mercancia_
    
    def importa_xls(self):
        wb_ = xlrd.open_workbook(self._archivo)
        ws_ = wb_.sheet_by_index(0)
        valores_ = []
        for indice_ in range(ws_.nrows):
            if ws_.row(indice_)[0] == self._campos[0]:
                pass
            else:
                mercancia_ = dict()
                for campo_ in range(len(self._campos)):
                    mercancia_.update({self._campos[campo_]:ws_.row(campo_)})
                    
                valores_.append(mercancia_)
        return valores_
    
    def verifica_formato(self, _campos):
        """
        Funcion que revisa que los campos necesarios se encuentren en el formato, de lo contrario
        devolvera un mensaje con el error.
        Recibe un diccionario.
        """
        
        columnas_ = ['House BL No.'
                     ,'Invoice Number'
                     ,'Material Code'
                     ,'Delivery Qty.'
                     ,'Qty Unit'
                     ,'Net Amount'
                     ,'Total Amount'
                     ,'Currency'
                     ,'Incoterms1'
                     ,'GS Vendor Name']
        respuesta_ = {'estatus':'Error'
                      ,'mensaje':''}
        mensaje_ = ''
        for campo_ in columnas_:
            if not _campos.has_key(campo_):
                mensaje_ = mensaje_ + 'No se encuentra el campo: %s'%campo_
        
        if mensaje_ == '' :
            respuesta_['estatus']='success'
        else:
            respuesta_['mensaje'] = mensaje_
        
        return respuesta_
    
    def importa_xlsx(self):
        "metodo para importar los archivos excel 2007 -"
        wb_ = load_workbook(filename=self._archivo, use_iterators=True)
        ws_ = wb_.worksheets[0]
        valores_ = []
        guias_distintas_ = []
        ban_ = 0
        estatus_ = None
        for fila_ in ws_.iter_rows():
            if ban_ == 0:
                
                for indice_ in range(len(fila_)):
                    #print fila_[indice_].internal_value
                    self._campos[fila_[indice_].internal_value] = indice_
                ban_ = 1
                estatus_ = self.verifica_formato(self._campos)
            else:
                if estatus_ :
                    mercancia_ = dict()
                    orden_ = sorted(self._campos.iteritems(),key=operator.itemgetter(1))
                    for campo_,indice_ in orden_:
                        if indice_:
                            #print indice_, campo_
                            try:
                                mercancia_.update({campo_.strip():fila_[indice_].internal_value})
                            except IndexError as e :
                                pass
                            # codigo para registrar los valores distintos
                            if campo_.strip()== 'House BL No.':
                                if fila_[indice_].internal_value in guias_distintas_:
                                    mercancia_['Repetida'] = 1
                                else:
                                    guias_distintas_.append(fila_[indice_].internal_value)
                                    mercancia_['Repetida'] = 0
                                mercancia_['fletes'] =0.00
                                mercancia_['embalajes'] =0.00
                                mercancia_['seguros'] =0.00
                                mercancia_['otros'] =0.00
                                mercancia_['moneda_guia'] =None
                            
                    valores_.append(mercancia_)
                else:
                    estatus_['mensaje'] = 'No se encontraron los encabezados de los campos.'
        
            
        
        return estatus_, valores_
    
class GLP:
    """
    Clase para el manejo de los archivos en excel del reporte GLP,
    obtenido del sistema GLP de Samsung, disponible mediante acceso web.
    Este reporte debe encontrarse en la primera hoja del libro.
    El cual contiene los siguientes campos, en ese orden:
        S/R No
        House B/L
        Master B/L
        Shipper
        Consignee
        Manufacturer
        Buyer
        Branch
        Ship To Party
        Partner	Partner(Carrier)
        Partner(3PL)
        Partner(Fwder)
        Partner(Trucker)
        P/O No
        Container No
        D/O No
        Sales Org.
        Division
        ShippingType
        Incoterms
        G/I Date
        ETD
        Loading Location(SA)
        S/R Create Date
        ETA(SA)
        ETA(Updated)
        Discharging Location
        Final Destination
        VSL/FLT
        Customs Broker
        Cleared Date
        Gross Weight
        Total Quantity
        Invoice Amount
        Invoice Currency
        Material Number
        Seller Plant
        Buyer Plant
        Consignee Storage Location
        B/L uploaded Date
        Billing
        Customs Status
        FTA Flag
    """
    
    _tipo = 0
    _archivo = ''
    _glp = []
    _campos = ['S/R No', 'House B/L', 'Master B/L', 'Shipper'
               , 'Consignee', 'Manufacturer', 'Buyer', 'Branch'
               , 'Ship To Party', 'Partner\tPartner(Carrier)'
               , 'Partner(3PL)', 'Partner(Fwder)', 'Partner(Trucker)'
               , 'P/O No', 'Container No', 'D/O No', 'Sales Org.'
               , 'Division', 'ShippingType', 'Incoterms', 'G/I Date'
               , 'ETD', 'Loading Location(SA)', 'S/R Create Date'
               , 'ETA(SA)', 'ETA(Updated)', 'Discharging Location'
               , 'Final Destination', 'VSL/FLT', 'Customs Broker'
               , 'Cleared Date', 'Gross Weight', 'Total Quantity'
               , 'Invoice Amount', 'Invoice Currency', 'Material Number'
               , 'Seller Plant', 'Buyer Plant', 'Consignee Storage Location'
               , 'B/L uploaded Date', 'Billing', 'Customs Status', 'FTA Flag']
    
    def __init__(self,_archivo):
        
        self._tipo = tipo_archivo(_archivo)
        if self._tipo != 0:
            self._archivo = _archivo
            self.carga_archivo()
        else:
            raise "El tipo de archivo o extesión, no coinciden con los de excel"
        return
    
    def carga_archivo(self):
        
        if self._tipo == 1:
            glp_ = self.importa_xls()
        elif self._tipo == 2 :
            glp_ = self.importa_xlsx()

        
        return glp_
    
    def importa_xls(self):
        wb_ = xlrd.open_workbook(self._archivo)
        
        
        return
    
    def importa_xlsx(self):
        """
        Metodo para importar el contenido del reporte a un objeto.
        """
        wb_ = load_workbook(filename=self._archivo, use_iterators= True)
        ws_ = wb_.worksheets[0]
        valores_ = []
        for fila_ in ws_.iter_rows():
            if fila_[0].internal_value == self._campos[0]:
                pass
            else:
                # defino un diccionario nuevo
                guia_ = dict()
                for indice_ in range(len(self._campos)):
                    #el cual se va llenando con los campos declarados y los valores
                    #obtenidos de la hoja en excel.
                    guia_.update({self._campos[indice_]:fila_[indice_].internal_value})
                valores_.append(guia_)
        
        return valores_

    